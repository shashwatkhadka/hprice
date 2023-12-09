import pandas as pd
from openpyxl import load_workbook

wb_p='Data/DataSet.xlsx'

df=pd.read_excel(wb_p)

#REMOVE NULL 
df.drop(columns=["PARKING","BUILDUP AREA"],inplace=True)
df.dropna(axis=0,inplace=True)

def conv2string(name):
    df[name]=df[name].astype(str)


def city_seperate(old,new):
    seperated=row[old].split(',')
    df.at[index,new]=seperated[-1]

def by_seperate(old,new):
    seperated=row[old].split()
    df.at[index,new]=seperated[0]

def land_seperate(old,new):
    
    seperated=row[old].split()
    if seperated[-1]=="katha":
        df.at[index,new]=float(seperated[0])*10.65
    else:    
        df.at[index,new]=(seperated[0])     

def road_seperate(old,new):
    seperated=row[old].split()
    if seperated[-1]=="Meter":
        df.at[index,new]=float(seperated[0])*3.28084
    else:    
        df.at[index,new]=seperated[0]

def price_seperate(old,new):
    seperate_array=row[old].split()
    if seperate_array[0]=="Rs.":
        numeric_part=seperate_array[1].replace(',','')
        if seperate_array[-1]=="Cr":
            df.at[index,new]=float(numeric_part)*10000000
        else:
            df.at[index,new]=float(numeric_part)
    else:
        df.drop(index,inplace=True)
   

conv2string("LAND AREA")
conv2string("ROAD ACCESS")
conv2string("BUILT YEAR")
conv2string("PRICE")
conv2string("FACING")

for index,row in df.iterrows():
    city_seperate("LOCATION","CITY")
    land_seperate("LAND AREA","LA_N")
    road_seperate("ROAD ACCESS","RA_N")
    by_seperate("BUILT YEAR","BY_N")
    price_seperate("PRICE","PRICE_N")

direction_priority={
    'East':1,
    'North':2,
    'North-East':3,
    'South-East':4,
    'North-West':5,
    'West':6,
    'South-West':7,
    'South':8
}
df['FACING_N']=df['FACING'].map(direction_priority)
 
workbook=load_workbook(wb_p)
workbook.create_sheet('FilteredData')

sheet=workbook['FilteredData']
filtercol=['CITY','LA_N','RA_N','BY_N','FLOOR','BEDROOM','BATHROOM','FACING_N','PRICE_N']

#create a new sheet and save data to it
for col_index,column in enumerate(filtercol,start=1):
    sheet.cell(row=1,column=col_index,value=column)
    for row_idx,value in enumerate(df[column],start=2):
        sheet.cell(row=row_idx,column=col_index,value=value)

workbook.save(wb_p)

