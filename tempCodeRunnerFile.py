
import pandas as pd
from openpyxl import load_workbook

df=pd.read_excel('Data/DataSet.xlsx')

#Earthquake Resistant
#Parking
#Drinking Water
#Parquet

for index,rows in df.iterrows():
    newval=rows["AMENITIES"].split(',')
    df.at[index,"AMENROW"]=newval

wb_p="Data/DataSet.xlsx"
workbook=load_workbook(wb_p)
workbook.create_sheet('amenities')

sheet=workbook['new']
filtercol=['AMENROW']

#create a new sheet and save data to it
for col_index,column in enumerate(filtercol,start=1):
    sheet.cell(row=1,column=col_index,value=column)
    for row_idx,value in enumerate(df[column],start=2):
        sheet.cell(row=row_idx,column=col_index,value=value)

workbook.save(wb_p)
